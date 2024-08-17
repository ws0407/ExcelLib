from matplotlib.ticker import AutoMinorLocator
from matplotlib.pyplot import MultipleLocator
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl

path = 'D:/Workspace/workspace_vscode_python/ExcelLib/数据/'

def get_data():
    """数据格式：年龄段+人数
    {
        '10~19': int,
        '20~29': int,
        ...,
        '90~99': int
    }
    """
    wb = openpyxl.load_workbook(path + 'results.xlsx')
    sheet1 = wb.active
    res = {f'{i}0~{i}9': 0 for i in range(1, 10)}
    # print(res)
    for row in range(77, 149):
        age = int(sheet1.cell(row, 1).value)
        res[f'{age // 10}0~{age // 10}9'] += int(sheet1.cell(row, 2).value)
    return res

def plot_1_subfigure():
    colors = ['brown', '#003399', '#C2DFFA', '#B6F0C9', '#FCD271' ,'#FF7A5A',
              '#f5587b', '#D87575', '#F05837', '#8AE1FC', '#0AAFF1', '#4ECDC4', '#C3a3E5']
    plt.figure(figsize=(8, 6))

    plt.rcParams['font.family'] = 'Cambria'
    plt.rcParams['font.size'] = 12
    plt.rcParams['axes.linewidth'] = 1
    # 设置图例标题大小
    plt.rcParams['legend.title_fontsize'] = 28

    data = get_data()
    # print(data)
    for key, i in zip(data.keys(), range(len(data.keys()))):
        p = plt.bar(key, data[key], color=colors[i])
        plt.bar_label(p, label_type='edge')
    plt.title("各年龄段balabala")
    plt.xlabel("年龄段", fontsize=18)
    plt.ylabel("例次", fontsize=18)
    plt.tick_params(axis='x', labelrotation=45)
    plt.gcf().subplots_adjust(top=0.95, bottom=0.23)
    plt.show()

exit()

from matplotlib.ticker import AutoMinorLocator
from matplotlib.pyplot import MultipleLocator
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl

path = 'D:/Workspace/workspace_vscode_python/ExcelLib/数据/'

def get_data():
    wb = openpyxl.load_workbook(path + 'results.xlsx')
    sheet1 = wb.active
    people = {}
    trade = {}
    for i in range(6):
        trade[sheet1.cell(i * 2 + 6, 2).value] = [int(sheet1.cell(i * 2 + 6, col).value) for col in range(3, 14)]
        people[sheet1.cell(i * 2 + 6, 2).value] = [int(sheet1.cell(i * 2 + 7, col).value) for col in range(3, 14)]
    return {'交易笔数': trade, '人数': people}


def plot_2_subfigure():
    res = get_data()
    # print(data)
    titles = ['']
    tick_titles = list(res.keys())
    diseases = list(res[tick_titles[0]].keys())
    keys_plot = diseases
    colors = ["#C72228", '#F98F34', '#0C4E9B', "#F5867F", "#FFBC80", "#6B98C4"]
    markers = ['o', 'x', 'v', '^', '.', 's']
    linestyles = ['-', ':', '-.', '--', '-', '-']

    x = ['2023-08', '2023-09', '2023-10', '2023-11', '2023-12', '2024-01', '2024-02', '2024-03', '2024-04', '2024-05', '2024-06']

    plt.rcParams['axes.linewidth'] = 1
    # 设置图例标题大小
    plt.rcParams['legend.title_fontsize'] = 10
    plt.rcParams["font.family"] = ["Times New Roman", "SimSun"]  # 使用支持的黑体中文字体
    # 全局字体
    plt.rcParams['font.size'] = 13
    # 标题的字体大小
    plt.rcParams['axes.titlesize'] = 13
    plt.rcParams["axes.unicode_minus"] = False  # 用来正常显示负数  "-"
    # 刻度字体大小
    plt.rcParams['xtick.labelsize'] = 10
    plt.rcParams['ytick.labelsize'] = 10
    # plt.rcParams['font.sans-serif'] = ['Fangsong']  # 用来正常显示中文标签

    # yticks(np.linspace(0.2, 1.0, 5, endpoint=True))

    fig, axes = plt.subplots(2, 2, figsize=(36, 36), dpi=120)
    for i in range(2):
        plt.delaxes(axes[1, i])

    plt.subplots_adjust(wspace=None, hspace=None)

    for i, disease in zip(range(6), diseases):
        for j in range(2):
            ax = axes[0, j]
            ax.plot(x, res[tick_titles[j]][disease], color=colors[i], label=disease, 
                         marker=markers[i], linestyle=linestyles[i])
            ax.plot([0], [0], color='w')
            ax.xaxis.set_minor_locator(MultipleLocator(1))
            ax.xaxis.set_major_locator(MultipleLocator(2))
            # 修改刻度属性
            ax.tick_params(which='major', axis='x', labelrotation=20, direction='in')
            ax.tick_params(which='minor', axis='x', direction='in')
            ax.tick_params(which='major', axis='y', direction='in')
            ax.set_xlabel(f'月份', labelpad=12)
            ax.set_title(f'({j+1}) {tick_titles[j]}', pad=10, loc='center') 
            # 添加网格
            ax.grid(which='major', ls='--', alpha=.8, lw=.8)
                      
            # 添加axis label
            ax.set_ylabel(tick_titles[j], labelpad=10)
        # 添加文本信息(标题)
        # axes[0, 0].set_title(' '*32 + '2023年8月-2024年6月某院门诊异地慢特病结算趋势图，（1）交易笔数；（2）人数。', pad=10, loc='left') 
        # 添加图例
        axes[0, 0].legend(fontsize=12, loc='upper left', title="", bbox_to_anchor=(0.37, -0.3), ncol=3, fancybox=True)
        
    # plt.savefig(path + '1.pdf')
    # plt.savefig('res/robustness_comparison_classes_new.png')
    # plt.savefig('res/malware-adv-big.png')
    plt.show()

plot_2_subfigure()


exit()


from matplotlib.ticker import AutoMinorLocator
from matplotlib.pyplot import MultipleLocator
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl

path = 'D:/Workspace/workspace_vscode_python/ExcelLib/数据/'

def get_data():
    res = {}
    wb = openpyxl.load_workbook(path + 'results.xlsx')
    sheet1 = wb.active
    sample = {}
    sample['交易笔数'] = [int(sheet1.cell(4, col).value) for col in range(3, 14)]
    sample['人数'] = [int(sheet1.cell(5, col).value) for col in range(3, 14)]
    res['总'] = sample
    return res

def plot_total_figure():
    res = get_data()
    # print(data)

    labels = list(res.keys())
    keys = list(res[labels[0]].keys())
    keys_plot = keys
    colors = ['#DE582B', '#1868B2']
    markers = ['o', '^']
    linestyles = ['-', '-.']

    x = ['2023-08', '2023-09', '2023-10', '2023-11', '2023-12', '2024-01', '2024-02', '2024-03', '2024-04', '2024-05', '2024-06']

    plt.rcParams['axes.linewidth'] = 1
    # 设置图例标题大小
    plt.rcParams['legend.title_fontsize'] = 10
    plt.rcParams["font.family"] = ["Times New Roman", "SimSun"]  # 使用支持的黑体中文字体
    # 全局字体
    plt.rcParams['font.size'] = 13
    # 标题的字体大小
    plt.rcParams['axes.titlesize'] = 13
    plt.rcParams["axes.unicode_minus"] = False  # 用来正常显示负数  "-"
    # 刻度字体大小
    plt.rcParams['xtick.labelsize'] = 10
    plt.rcParams['ytick.labelsize'] = 10
    # plt.rcParams['font.sans-serif'] = ['Fangsong']  # 用来正常显示中文标签

    # yticks(np.linspace(0.2, 1.0, 5, endpoint=True))

    fig, axes = plt.subplots(2, 2, figsize=(48, 48), dpi=110)
    # plt.delaxes(axes[2, 1])

    plt.subplots_adjust(wspace=None, hspace=0.6)

    for i, label in zip(range(1), labels):
        ax = axes[int(i / 3), int(i % 3)]
        for p, c, m, ls, k_p in zip(keys, colors, markers, linestyles, keys_plot):
            ax.plot(x[:len(res[label][p])], res[label][p][:], color=c, label=k_p, marker=m, linestyle=ls)
        ax.plot([0], [0], color='w')
        # 修改次刻度
        # yminorLocator = MultipleLocator(1)  # 将此y轴次刻度标签设置为0.1的倍数
        xminorLocator = MultipleLocator(1)
        xmajorLocator = MultipleLocator(2)
        # ymajorLocator = MultipleLocator(100)
        # ax.yaxis.set_minor_locator(yminorLocator)
        ax.xaxis.set_minor_locator(xminorLocator)
        # ax.yaxis.set_major_locator(ymajorLocator)
        ax.xaxis.set_major_locator(xmajorLocator)

        if i == 5:
            ax.yaxis.set_minor_locator(MultipleLocator(1))
            ax.yaxis.set_major_locator(MultipleLocator(1))
        # 修改刻度属性
        ax.tick_params(which='major', axis='x', labelrotation=20)
        ax.tick_params(which='major', axis='y')
        
        # 添加axis label
        if i % 3 == 0:
            ax.set_ylabel('人数/交易笔数', labelpad=10 if i == 0 else 15)
        if i == 0:
            ax.set_xlabel('月份', labelpad=12)
        # 添加网格
        ax.grid(which='major', ls='--', alpha=.8, lw=.8)
        # 添加图例
        if ax == axes[0, 0]:
            ax.legend(fontsize=12, loc='upper left', title="", bbox_to_anchor=(0.25, -0.3), ncol=2, fancybox=True)
        # 添加文本信息(标题)
        ax.set_title('2023年8月-2024年6月某院门诊异地慢特病\n结算交易笔数和人数趋势图', pad=10, loc='center')
    # plt.savefig(path + '1.pdf')
    # plt.savefig('res/robustness_comparison_classes_new.png')
    # plt.savefig('res/malware-adv-big.png')
    plt.show()

plot_total_figure()



exit()

from matplotlib.ticker import AutoMinorLocator
from matplotlib.pyplot import MultipleLocator
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time
import openpyxl

path = 'D:/Workspace/workspace_vscode_python/ExcelLib/数据/'


def get_data():
    res = {}
    wb = openpyxl.load_workbook(path + 'results.xlsx')
    sheet1 = wb.active

    for i in range(6):
        sample = {}
        sample['人数'] = [int(sheet1.cell(i * 2 + 6, col).value) for col in range(3, 14)]
        sample['交易笔数'] = [int(sheet1.cell(i * 2 + 7, col).value) for col in range(3, 14)]
        res[sheet1.cell(i * 2 + 6, 2).value] = sample
    return res


def plot_6_subfigure():
    res = get_data()
    # print(data)

    labels = list(res.keys())
    keys = list(res[labels[0]].keys())
    keys_plot = keys
    colors = ['#DE582B', '#1868B2']
    markers = ['o', 'x']
    linestyles = ['-', '-.']

    x = ['2023-08', '2023-09', '2023-10', '2023-11', '2023-12', '2024-01', '2024-02', '2024-03', '2024-04', '2024-05', '2024-06']

    plt.rcParams['axes.linewidth'] = 1
    # 设置图例标题大小
    plt.rcParams['legend.title_fontsize'] = 10
    plt.rcParams["font.family"] = ["Times New Roman", "SimSun"]  # 使用支持的黑体中文字体
    # 全局字体
    plt.rcParams['font.size'] = 13
    # 标题的字体大小
    plt.rcParams['axes.titlesize'] = 13
    plt.rcParams["axes.unicode_minus"] = False  # 用来正常显示负数  "-"
    # 刻度字体大小
    plt.rcParams['xtick.labelsize'] = 10
    plt.rcParams['ytick.labelsize'] = 10
    # plt.rcParams['font.sans-serif'] = ['Fangsong']  # 用来正常显示中文标签

    # yticks(np.linspace(0.2, 1.0, 5, endpoint=True))

    fig, axes = plt.subplots(3, 3, figsize=(36, 36), dpi=120)
    for i in range(3):
        plt.delaxes(axes[2, i])
    # plt.delaxes(axes[2, 1])

    plt.subplots_adjust(wspace=None, hspace=0.6)

    for i, label in zip(range(6), labels):
        ax = axes[int(i / 3), int(i % 3)]
        for p, c, m, ls, k_p in zip(keys, colors, markers, linestyles, keys_plot):
            ax.plot(x[:len(res[label][p])], res[label][p][:], color=c, label=k_p, marker=m, linestyle=ls)
        ax.plot([0], [0], color='w')
        # 修改次刻度
        # yminorLocator = MultipleLocator(1)  # 将此y轴次刻度标签设置为0.1的倍数
        xminorLocator = MultipleLocator(1)
        xmajorLocator = MultipleLocator(2)
        # ymajorLocator = MultipleLocator(100)
        # ax.yaxis.set_minor_locator(yminorLocator)
        ax.xaxis.set_minor_locator(xminorLocator)
        # ax.yaxis.set_major_locator(ymajorLocator)
        ax.xaxis.set_major_locator(xmajorLocator)

        if i == 5:
            ax.yaxis.set_minor_locator(MultipleLocator(1))
            ax.yaxis.set_major_locator(MultipleLocator(1))
        # 修改刻度属性
        ax.tick_params(which='major', axis='x', labelrotation=20, direction='in', top='on', right="on")
        ax.tick_params(which='major', axis='y', direction='in', top='on', right="on")
        
        # 添加axis label
        if i % 3 == 0:
            ax.set_ylabel('人数/交易笔数', labelpad=10 if i == 0 else 15)
        if i == 4:
            ax.set_xlabel('月份', labelpad=12)
        # 添加网格
        ax.grid(which='major', ls='--', alpha=.8, lw=.8)
        # 添加图例
        if ax == axes[1, 1]:
            ax.legend(fontsize=12, loc='upper left', title="", bbox_to_anchor=(0.04, -0.55), ncol=2, fancybox=True)
        # 添加文本信息(标题)
        ax.set_title(f'({i+1}) ' + label, pad=10, loc='center')
    # plt.savefig(path + '1.pdf')
    # plt.savefig('res/robustness_comparison_classes_new.png')
    # plt.savefig('res/malware-adv-big.png')
    plt.show()

plot_6_subfigure()


exit()

filepath = "D:/Workspace/workspace_vscode_python/ExcelLib/数据/"
filename = "2023.8-2024.6合并数据_split.csv"
# filename = '120109194004221029-总.csv'
filename2 = "异地门诊合并_1-2_split.csv"
# filename2 = '120109194004221029-门诊.csv'

s_t = time.time()
df = pd.read_csv(filepath + filename)
df2 = pd.read_csv(filepath + filename2)
print("读取数据耗时:", time.time() - s_t, "秒")

df = df.rename(columns={"身份证号": "身份证号码"})

# 根据日期，身份证号码，金额求两个dataframe的交集
df3 = pd.merge(df, df2, how="inner", on=["日期", "身份证号码", "总金额"])

cols = sorted([_ for _ in df3.columns])
print(cols)

df3 = df3.reindex(columns=cols)

df3.to_csv(
    filepath + "2023.8-2024.6合并数据_异地门诊合并_merge.csv",
    index=False,
    encoding="utf-8",
)
"""
合并结果：82715条数据
原总数据：179208条数据
门诊数据：118941条数据
"""
exit()

filepath = "D:/Workspace/workspace_vscode_python/ExcelLib/数据/"
filename = "2023.8-2024.6合并数据_new.csv"
filename2 = "异地门诊合并_1-2.csv"

s_t = time.time()
df = pd.read_csv(filepath + filename)
df2 = pd.read_csv(filepath + filename2)
print("读取数据耗时:", time.time() - s_t, "秒")

df[["日期", "时间"]] = df["交易日期"].str.split(" ", expand=True)
df2[["日期", "时间"]] = df2["收费日期"].str.split(" ", expand=True)

df["日期"] = pd.to_datetime(df["日期"])
df2["日期"] = pd.to_datetime(df2["日期"])

df.drop(["交易日期"], axis=1, inplace=True)
df2.drop(["收费日期"], axis=1, inplace=True)

df.to_csv(filepath + "2023.8-2024.6合并数据_split.csv", index=False, encoding="utf-8")
df2.to_csv(filepath + "异地门诊合并_1-2_split.csv", index=False, encoding="utf-8")

exit()

filepath = "D:/Workspace/workspace_vscode_python/ExcelLib/数据/"
filename = "2023.8-2024.6合并数据_new.csv"
filename2 = "异地门诊合并_1-2.csv"

s_t = time.time()
df = pd.read_csv(filepath + filename)
df2 = pd.read_csv(filepath + filename2)
print("读取数据耗时:", time.time() - s_t, "秒")

df = df[df["身份证号"] == "120109194004221029"]
df.to_csv(filepath + "120109194004221029-总.csv", encoding="gbk")

df2 = df2[df2["身份证号码"] == "120109194004221029"]
df2.to_csv(filepath + "120109194004221029-门诊.csv", encoding="gbk")

exit()

s_t = time.time()
df = pd.read_csv(filepath + filename)
df2 = pd.read_csv(filepath + filename2)
df3 = pd.read_csv(filepath + filename3)
print("读取数据耗时:", time.time() - s_t, "秒")

print("2023.8-2024.6合并数据_new.csv: \n", df["医疗类别"].value_counts())
print("\n\n异地门诊合并_1-2.csv: \n", df2["医疗类别"].value_counts())
print("\n\n异地挂号合并_1-2.csv: \n", df3["医疗类别"].value_counts())

exit()

s_t = time.time()
filepath = (
    "D:/Workspace/workspace_vscode_python/ExcelLib/数据/" "2023.8-2024.6合并数据.csv"
)

df = pd.read_csv(filepath)
print("读取数据耗时:", time.time() - s_t, "秒")

df = df.drop(df[df["险种类别"] == "合计"].index)

df.to_csv(filepath[:-4] + "_new.csv")

exit()

s_t = time.time()
filepath = (
    "D:/Workspace/workspace_vscode_python/ExcelLib/数据/"
    "异地门诊患者结算数据20230801-20240630新.xls"
)

sheet_names = [
    "门诊20230801-20240131",
    "门诊20240201-0630",
    "挂号20230801-1231",
    "挂号20240101-0430",
    "挂号20240501-0630",
]

# df = pd.read_excel(filepath, dtype=str, sheet_name=None)
df1 = pd.read_excel(filepath, dtype=str, sheet_name=sheet_names[2])
print("读取数据耗时:", time.time() - s_t, "秒")
df2 = pd.read_excel(filepath, dtype=str, sheet_name=sheet_names[3])
print("读取数据耗时:", time.time() - s_t, "秒")
df3 = pd.read_excel(filepath, dtype=str, sheet_name=sheet_names[4])
print("读取数据耗时:", time.time() - s_t, "秒")
# 合并两个数据表
df = pd.concat([df1, df2, df3], ignore_index=True)
print("合并完成，正在保存...")

df.to_csv(filepath[:-4] + "_3-5.csv")

exit()

# 测试数据
# df = pd.DataFrame({
#         'Index': list(range(1, 8)),
#         '交易流水号': ['one', 'one', 'two', 'BBB', 'four', 'three', 'four'],
#         '原交易流水号': [np.nan, np.nan, np.nan, 'two', 'three', np.nan, np.nan],
#     })

nums = df[df["原交易流水号"].notnull()]["原交易流水号"].tolist()
# print(df.tail())
print("共删除", len(nums) * 2, "条数据")
df = df.drop(df[df["交易流水号"].isin(nums)].index)
df = df.drop(df[df["原交易流水号"].notnull()].index)

# 一行
# df = df.drop(df[(df['交易流水号'].isin(df[df['原交易流水号'].notnull()]['原交易流水号'].tolist())) /
#                  | (df['原交易流水号'].notnull())].index)

# print(df.tail())

# 保存数据，gbk是中文编码
df.to_csv(
    r"D:/Workspace/workspace_vscode_python/ExcelLib/数据/2023.8-2024.6合并数据_new.csv",
    encoding="gbk",
)
